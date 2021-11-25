import datetime
import os

from applications.enums import ApplicationStatus, BenefitType
from applications.tests.factories import ApplicationFactory
from calculator.models import Calculation, STATE_AID_MAX_PERCENTAGE_CHOICES
from calculator.tests.factories import PaySubsidyFactory
from common.utils import nested_setattr, to_decimal
from helsinkibenefit.tests.conftest import *  # noqa
from openpyxl import load_workbook


class CaseNotFound(Exception):
    pass


def make_bool_fi(value):
    if isinstance(value, str):
        if value.lower() in ["kyllä", "k"]:
            return True
        elif value.lower() in ["ei", "e"]:
            return False
    elif isinstance(value, bool):
        return value
    else:
        raise ValueError(f"Invalid value: {value}")


MAX_TEST_ROW = 100
MAX_TEST_COLUMN = 5  # complex cases not tested yet


class ExpectedResults:
    def __init__(self, **attribute_name_values):
        for name, value in attribute_name_values.items():
            setattr(self, name, value)


class ExcelTestCase:
    def __init__(self, worksheet, column_idx):
        self.worksheet = worksheet
        self.column_idx = column_idx

    BENEFIT_TYPE_MAP = {
        "Palkan Helsinki-lisä": BenefitType.SALARY_BENEFIT,
        "Työllistämisen Helsinki-lisä": BenefitType.EMPLOYMENT_BENEFIT,
    }


# unique object
sentinel = object()


class SalaryBenefitExcelTest(ExcelTestCase):
    def _setup(self):
        self._setup_expected_results()
        self._setup_db_objects()
        self._load_values_from_excel()
        self._save_initial_state()

    def _setup_expected_results(self):
        # actual expected values loaded from Excel
        self.expected_results = ExpectedResults(
            calculated_benefit_amount=sentinel,
            salary_costs=sentinel,
            state_aid_max_monthly_eur=sentinel,
            time_range_1=ExpectedResults(
                start_date=sentinel,
                end_date=sentinel,
                duration=sentinel,
                pay_subsidy_monthly=sentinel,
                monthly_amount=sentinel,
                total_amount=sentinel,
            ),
            time_range_2=ExpectedResults(
                start_date=sentinel,
                end_date=sentinel,
                duration=sentinel,
                pay_subsidy_monthly=sentinel,
                monthly_amount=sentinel,
                total_amount=sentinel,
            ),
        )

    def _setup_db_objects(self):

        self.application = ApplicationFactory()
        self.application.status = ApplicationStatus.RECEIVED
        self.application.save()

        self.application.calculation = Calculation(
            application=self.application,
            monthly_pay=self.application.employee.monthly_pay,
            vacation_money=self.application.employee.vacation_money,
            other_expenses=self.application.employee.other_expenses,
            start_date=self.application.start_date,
            end_date=self.application.end_date,
            state_aid_max_percentage=STATE_AID_MAX_PERCENTAGE_CHOICES[0][0],
            calculated_benefit_amount=0,
            override_benefit_amount=None,
        )

        self.application.pay_subsidy_1 = PaySubsidyFactory()
        self.application.pay_subsidy_1.pay_subsidy_percent = None
        self.application.pay_subsidy_2 = PaySubsidyFactory()
        self.application.pay_subsidy_2.pay_subsidy_percent = None

    value_conversion_table = {
        "Palkan Helsinki-lisä": BenefitType.SALARY_BENEFIT,
        "Työllistämisen Helsinki-lisä": BenefitType.EMPLOYMENT_BENEFIT,
        "kyllä": True,
        "ei": False,
    }

    def _load_values_from_excel(self):
        for row_idx in range(1, MAX_TEST_ROW):
            target = self.worksheet.cell(row_idx, 1).value  # dotted object
            if not target:
                continue
            value = self.worksheet.cell(row_idx, self.column_idx).value
            if value in self.value_conversion_table:
                value = self.value_conversion_table[value]
            elif target.endswith("_date"):
                value = self.convert_date(value)
            elif target.endswith("percent") or target.endswith("percentage"):
                if value is not None:
                    value = to_decimal(value * 100, decimal_places=2)
            elif isinstance(target, float):
                if value is not None:
                    value = to_decimal(value, decimal_places=2)

            print(f"{target}={value} ({type(value)}")
            nested_setattr(self, target, value)

    def convert_date(self, value):
        if value in [None, ""]:
            value = None
        elif isinstance(value, datetime.datetime):
            value = value.date()
        elif isinstance(value, str):
            value = datetime.datetime.strptime(value, "%d/%m/%Y")
        else:
            raise ValueError(f"Unexpected {type(value)} value: {value}")
        return value

    def _save_initial_state(self):
        self.application.save()
        self.application.calculation.save()

        for pay_subsidy in [
            self.application.pay_subsidy_1,
            self.application.pay_subsidy_2,
        ]:
            if pay_subsidy.pay_subsidy_percent is not None:
                # only add the pay subsidy to application if the subsidy is defined
                # in the Excel testcase
                pay_subsidy.application = self.application
                pay_subsidy.save()

        self.application.calculation.start_date = (
            self.expected_results.time_range_1.start_date
        )
        self.application.calculation.end_date = (
            self.expected_results.time_range_2.end_date
            or self.expected_results.time_range_1.end_date
        )
        self.application.calculation.save()

    def run_test(self):
        self._setup()
        self.application.calculation.init_calculator()
        self.application.calculation.calculate()
        self._verify_results()

    def _verify_results(self):
        for row in self.application.calculation.rows.all():
            print(row)
        assert (
            self.application.calculation.calculated_benefit_amount
            == self.expected_results.calculated_benefit_amount
        )


class EmployeeBenefitExcelTest(SalaryBenefitExcelTest):
    def _setup_expected_results(self):
        # actual expected values loaded from Excel
        self.expected_results = ExpectedResults(
            calculated_benefit_amount=sentinel,
            salary_costs=sentinel,
            state_aid_max_monthly_eur=sentinel,
            start_date=sentinel,
            end_date=sentinel,
            duration=sentinel,
            monthly_amount=sentinel,
            total_amount=sentinel,
        )

    def _setup_db_objects(self):

        self.application = ApplicationFactory()
        self.application.status = ApplicationStatus.RECEIVED
        self.application.save()

        self.application.calculation = Calculation(
            application=self.application,
            monthly_pay=self.application.employee.monthly_pay,
            vacation_money=self.application.employee.vacation_money,
            other_expenses=self.application.employee.other_expenses,
            start_date=self.application.start_date,
            end_date=self.application.end_date,
            state_aid_max_percentage=STATE_AID_MAX_PERCENTAGE_CHOICES[0][0],
            calculated_benefit_amount=0,
            override_benefit_amount=None,
        )

        self.application.pay_subsidy = PaySubsidyFactory()
        self.application.pay_subsidy.pay_subsidy_percent = None

    def _save_initial_state(self):
        self.application.save()
        self.application.calculation.save()

        self.application.calculation.start_date = self.expected_results.start_date
        self.application.calculation.end_date = self.expected_results.end_date
        self.application.calculation.save()

    def run_test(self):
        self._setup()
        self.application.calculation.init_calculator()
        self.application.calculation.calculate()
        self._verify_results()

    def _verify_results(self):
        for row in self.application.calculation.rows.all():
            print(row)
        assert (
            self.application.calculation.calculated_benefit_amount
            == self.expected_results.calculated_benefit_amount
        )


FIRST_TEST_COLUMN = 3


def test_salary_benefit_cases_from_excel(request, api_client):

    excel_file_name = os.path.join(
        request.fspath.dirname, "Helsinki-lisä laskurin testitapaukset.xlsx"
    )
    wb = load_workbook(filename=excel_file_name, data_only=True)  # do not load formulas
    salary_benefit_tests = wb["Palkan Helsinki-lisä"]
    for col_idx in range(FIRST_TEST_COLUMN, MAX_TEST_COLUMN):
        try:
            test = SalaryBenefitExcelTest(salary_benefit_tests, col_idx)
            # add_attachments_to_application(request, application)
        except CaseNotFound:  # no tests
            break
        else:
            test.run_test()


def test_employee_benefit_cases_from_excel(request, api_client):

    excel_file_name = os.path.join(
        request.fspath.dirname, "Helsinki-lisä laskurin testitapaukset.xlsx"
    )
    wb = load_workbook(filename=excel_file_name, data_only=True)  # do not load formulas
    salary_benefit_tests = wb["Työllistämisen Helsinki-lisä"]
    for col_idx in range(FIRST_TEST_COLUMN, MAX_TEST_COLUMN):
        try:
            test = EmployeeBenefitExcelTest(salary_benefit_tests, col_idx)
            # add_attachments_to_application(request, application)
        except CaseNotFound:  # no tests
            break
        else:
            test.run_test()
