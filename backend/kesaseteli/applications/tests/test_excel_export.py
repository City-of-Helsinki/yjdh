import random
from datetime import date, datetime, timedelta
from decimal import Decimal
from io import BytesIO
from typing import List
from urllib.parse import parse_qs, urlparse

import openpyxl
import pytest
from auditlog.models import LogEntry
from django.contrib.contenttypes.models import ContentType
from django.http import StreamingHttpResponse
from django.shortcuts import reverse
from django.test import override_settings
from django.utils import translation
from django.utils.timezone import localdate
from django.utils.translation import gettext_lazy as _
from freezegun import freeze_time
from rest_framework import status

from applications.api.handler_excel_views import YouthApplicationExcelExportViewSet
from applications.employer_excel_export import (
    EmployerExcelExportErrorCode,
    get_excel_download_error_message,
)
from applications.enums import (
    AdditionalInfoUserReason,
    EmployerApplicationStatus,
    ExcelColumns,
    VtjTestCase,
    YouthApplicationStatus,
)
from applications.exporters.excel_exporter import (
    APPLICATION_LANGUAGE_FIELD_TITLE,
    EMPLOYMENT_END_DATE_FIELD_TITLE,
    EMPLOYMENT_START_DATE_FIELD_TITLE,
    ExcelField,
    FIELDS,
    get_attachment_uri,
    get_exportable_fields,
    get_reporting_columns,
    get_talpa_columns,
    handle_special_cases,
    HIRED_WITHOUT_VOUCHER_ASSESSMENT_FIELD_TITLE,
    INVOICER_EMAIL_FIELD_TITLE,
    INVOICER_NAME_FIELD_TITLE,
    INVOICER_PHONE_NUMBER_FIELD_TITLE,
    ORDER_FIELD_TITLE,
    RECEIVED_DATE_FIELD_TITLE,
    REMOVABLE_REPORTING_FIELD_TITLES,
    REMOVABLE_TALPA_FIELD_TITLES,
    SALARY_PAID_FIELD_TITLE,
    SPECIAL_CASE_FIELD_TITLE,
    SUM_FIELD_TITLE,
    TALPA_END_FIELD_TITLES,
    VOUCHER_NUMBER_FIELD_TITLE,
    WORK_HOURS_FIELD_TITLE,
)
from applications.models import EmployerSummerVoucher, YouthApplication
from applications.tests.test_models import create_test_employer_summer_vouchers
from common.tests.factories import (
    ActiveVtjTestCaseYouthApplicationFactory,
    ActiveYouthApplicationFactory,
    EmployerApplicationFactory,
    EmployerSummerVoucherFactory,
    InactiveYouthApplicationFactory,
    YouthApplicationFactory,
)
from common.urls import handler_403_url
from common.utils import getattr_nested
from shared.common.tests.utils import utc_datetime


def excel_download_url():
    return reverse("excel-download")


# Shared with youth export inline empty response (handler_excel_views.py).
NO_APPLICATIONS_FINNISH = "Hakemuksia ei löytynyt."


def assert_redirects_to_landing_page_with_error(
    response,
    expected_error_code: EmployerExcelExportErrorCode,
):
    """Assert the export endpoint redirected with an allowlisted ``error`` code."""
    assert response.status_code == status.HTTP_302_FOUND
    assert urlparse(response.url).path == excel_download_url()
    query = parse_qs(urlparse(response.url).query)
    assert query.get("error") == [expected_error_code.value], query


def assert_landing_page_shows_export_error(
    staff_client,
    redirect_response,
    expected_error_code: EmployerExcelExportErrorCode,
):
    """Follow the export redirect and verify the landing page shows the mapped alert."""
    assert_redirects_to_landing_page_with_error(redirect_response, expected_error_code)
    landing_response = staff_client.get(redirect_response.url)
    assert landing_response.status_code == status.HTTP_200_OK
    assert (
        get_excel_download_error_message(expected_error_code)
        in landing_response.content.decode()
    )


def employer_excel_export_url(
    export_kind: str,
    columns: str = ExcelColumns.REPORTING.value,
) -> str:
    return reverse(
        "employer-excel-export",
        kwargs={"export_kind": export_kind, "columns": columns},
    )


def youth_excel_download_url():
    return reverse("youth-excel-download")


def get_field_titles(fields: List[ExcelField]) -> List[str]:
    return [field.title for field in fields]


def check_removable_field_titles(removable_field_titles):
    assert len(removable_field_titles) == len(set(removable_field_titles))
    assert set(removable_field_titles) <= set(get_field_titles(FIELDS))


@pytest.mark.django_db
@override_settings(NEXT_PUBLIC_MOCK_FLAG=False)
@pytest.mark.parametrize("http_method", ["delete", "patch", "post", "put"])
def test_excel_view_unallowed_methods(request, staff_client, http_method):
    client_http_method_func = getattr(staff_client, http_method)
    assert callable(client_http_method_func)
    response = client_http_method_func(
        employer_excel_export_url("annual", ExcelColumns.REPORTING.value)
    )
    assert response.status_code == status.HTTP_405_METHOD_NOT_ALLOWED


@pytest.mark.django_db
@override_settings(NEXT_PUBLIC_MOCK_FLAG=False)
def test_excel_view_get_with_authenticated_user(staff_client):
    response = staff_client.get(excel_download_url())
    assert response.status_code == 200


@pytest.mark.django_db
@override_settings(NEXT_PUBLIC_MOCK_FLAG=False)
def test_excel_view_get_with_unauthenticated_user(user_client):
    response = user_client.get(excel_download_url())
    assert response.status_code == 302
    assert response.url == handler_403_url()


@pytest.mark.django_db
@override_settings(NEXT_PUBLIC_MOCK_FLAG=False)
def test_excel_view_download_unhandled(
    staff_client, submitted_summer_voucher, submitted_employment_contract_attachment
):
    submitted_summer_voucher.application.status = EmployerApplicationStatus.SUBMITTED
    submitted_summer_voucher.application.save()

    response = staff_client.get(
        employer_excel_export_url("unhandled", ExcelColumns.REPORTING.value)
    )

    assert response.status_code == 200
    submitted_summer_voucher.refresh_from_db()
    assert submitted_summer_voucher.is_exported is True
    # Cannot decode an xlsx file
    with pytest.raises(UnicodeDecodeError):
        response.getvalue().decode()


@pytest.mark.django_db
@override_settings(NEXT_PUBLIC_MOCK_FLAG=False)
def test_excel_view_download_no_unhandled_applications(staff_client):
    """Empty unhandled export redirects to the landing page with ``no_unhandled``."""
    response = staff_client.get(
        employer_excel_export_url("unhandled", ExcelColumns.REPORTING.value)
    )

    assert_landing_page_shows_export_error(
        staff_client, response, EmployerExcelExportErrorCode.NO_UNHANDLED
    )


@pytest.mark.django_db
@override_settings(NEXT_PUBLIC_MOCK_FLAG=False)
@pytest.mark.parametrize("columns", ExcelColumns.values)
def test_excel_view_download_no_annual_applications(staff_client, columns):
    """Empty annual export redirects to the landing page with ``no_applications``."""
    # Create draft applications with/without voucher, these should not be returned
    EmployerSummerVoucherFactory(
        application=EmployerApplicationFactory(status=EmployerApplicationStatus.DRAFT)
    )
    EmployerApplicationFactory(status=EmployerApplicationStatus.DRAFT)

    response = staff_client.get(employer_excel_export_url("annual", columns))

    assert_landing_page_shows_export_error(
        staff_client, response, EmployerExcelExportErrorCode.NO_APPLICATIONS
    )


@pytest.mark.django_db
@override_settings(NEXT_PUBLIC_MOCK_FLAG=False)
def test_youth_excel_download_no_youth_applications(staff_client):
    """Youth empty export returns 200 with inline text, not a landing-page redirect."""
    response = staff_client.get(youth_excel_download_url())
    assert response.status_code == 200
    assert NO_APPLICATIONS_FINNISH in response.content.decode()


@pytest.mark.django_db
def test_youth_excel_download_writes_audit_log(staff_client, youth_application):
    old_audit_log_entry_count = LogEntry.objects.count()
    frozen_datetime = utc_datetime(2025, 10, 15)
    with freeze_time(frozen_datetime):
        response = staff_client.get(youth_excel_download_url())

    assert LogEntry.objects.count() == old_audit_log_entry_count + 1
    audit_event = LogEntry.objects.filter(
        action=LogEntry.Action.ACCESS,
        timestamp=frozen_datetime,
    ).first()
    assert audit_event.actor_id == response.wsgi_request.user.pk
    assert audit_event.actor_email == response.wsgi_request.user.email
    assert audit_event.action == LogEntry.Action.ACCESS
    assert audit_event.object_pk == ""
    assert audit_event.content_type == ContentType.objects.get_for_model(
        YouthApplication
    )
    assert audit_event.additional_data == {
        "is_sent": False,
        "request_path": youth_excel_download_url(),
        "method": "YouthApplicationExcelExportViewSet.list",
    }
    assert audit_event.timestamp == frozen_datetime


@pytest.mark.django_db
@override_settings(NEXT_PUBLIC_MOCK_FLAG=False)
@pytest.mark.parametrize("columns", ExcelColumns.values)
@pytest.mark.parametrize("download_type", ["unhandled", "annual", "annual-previous"])
def test_excel_download_writes_audit_log(staff_client, columns, download_type):
    """
    Test that audit log is written when downloading employer summer voucher
    Excel files.
    """
    create_test_employer_summer_vouchers(year=2021)
    download_url = employer_excel_export_url(download_type, columns)

    old_audit_log_entry_count = LogEntry.objects.count()
    frozen_datetime = utc_datetime(2021, 12, 31)
    with freeze_time(frozen_datetime):
        response = staff_client.get(download_url)

    assert LogEntry.objects.count() == old_audit_log_entry_count + 1
    audit_event = LogEntry.objects.filter(
        action=LogEntry.Action.ACCESS,
        timestamp=frozen_datetime,
    ).first()
    assert audit_event.action == LogEntry.Action.ACCESS
    assert audit_event.object_pk == ""  # Not a single object, but a list
    assert audit_event.content_type == ContentType.objects.get_for_model(
        EmployerSummerVoucher
    )
    assert audit_event.additional_data == {
        "is_sent": False,
        "request_path": download_url,
        "method": "EmployerApplicationExcelExportView.get",
        "parameters": {
            "columns": columns,
            "export_kind": download_type,
        },
    }
    assert audit_event.timestamp == frozen_datetime
    assert audit_event.actor_id == response.wsgi_request.user.id
    assert audit_event.actor_email == response.wsgi_request.user.email


@pytest.mark.django_db
@override_settings(NEXT_PUBLIC_MOCK_FLAG=False)
def test_excel_view_download_annual(
    staff_client, submitted_summer_voucher, submitted_employment_contract_attachment
):
    submitted_summer_voucher.application.status = EmployerApplicationStatus.SUBMITTED
    submitted_summer_voucher.application.save()

    response = staff_client.get(
        employer_excel_export_url("annual", ExcelColumns.REPORTING.value)
    )

    assert response.status_code == 200
    submitted_summer_voucher.refresh_from_db()
    assert submitted_summer_voucher.is_exported is False
    # Cannot decode an xlsx file
    with pytest.raises(UnicodeDecodeError):
        response.getvalue().decode()


@pytest.mark.django_db
@override_settings(NEXT_PUBLIC_MOCK_FLAG=False)
@pytest.mark.parametrize(
    "download_url,expected_output_excel_fields",
    [
        (
            employer_excel_export_url(download, columns),
            get_exportable_fields(columns),
        )
        for columns in ExcelColumns.values
        for download in ["unhandled", "annual"]
    ],
)
def test_excel_view_download_content(  # noqa: C901
    staff_client,
    download_url,
    expected_output_excel_fields: List[ExcelField],
):
    def employer_summer_voucher_sorting_key(voucher: EmployerSummerVoucher):
        # Sorting key should be the same as what is used to order by queryset results
        # in Excel download, see EmployerExcelExportService.base_queryset
        return voucher.submitted_at, voucher.created_at, voucher.pk

    vouchers: List[EmployerSummerVoucher] = sorted(
        create_test_employer_summer_vouchers(year=2021),
        key=employer_summer_voucher_sorting_key,
    )

    with freeze_time(datetime(2021, 12, 31)):
        response = staff_client.get(download_url)

    assert isinstance(response, StreamingHttpResponse)

    workbook = openpyxl.load_workbook(filename=BytesIO(response.getvalue()))
    active_worksheet = workbook.active

    rows_generator = active_worksheet.rows
    header_row = next(rows_generator)
    data_rows = [next(rows_generator) for _ in range(len(vouchers))]
    with pytest.raises(StopIteration):
        next(rows_generator)

    expected_output_field_titles = get_field_titles(expected_output_excel_fields)
    output_field_titles = [
        "" if column.value is None else column.value  # Export changes "" to None
        for column in header_row
    ]

    # Check that output header is correct
    assert output_field_titles == expected_output_field_titles

    # Check that output data is correct
    for row_number, (data_row, voucher) in enumerate(zip(data_rows, vouchers), start=1):
        for output_column, excel_field in zip(data_row, expected_output_excel_fields):
            if excel_field.title == ORDER_FIELD_TITLE:
                assert output_column.value == row_number, "Incorrect output sorting"
            elif excel_field.title == RECEIVED_DATE_FIELD_TITLE:
                assert (
                    output_column.value
                    == voucher.submitted_at.astimezone().strftime("%d/%m/%Y")
                )
            elif excel_field.title == APPLICATION_LANGUAGE_FIELD_TITLE:
                assert output_column.value == voucher.application.get_language_display()
            elif excel_field.title == SPECIAL_CASE_FIELD_TITLE:
                assert output_column.value == voucher.get_target_group_display()
            elif excel_field.title == EMPLOYMENT_START_DATE_FIELD_TITLE:
                assert output_column.value == voucher.employment_start_date.strftime(
                    "%d.%m.%Y"
                )
            elif excel_field.title == EMPLOYMENT_END_DATE_FIELD_TITLE:
                assert output_column.value == voucher.employment_end_date.strftime(
                    "%d.%m.%Y"
                )
            elif excel_field.title == HIRED_WITHOUT_VOUCHER_ASSESSMENT_FIELD_TITLE:
                assert (
                    output_column.value
                    == voucher.get_hired_without_voucher_assessment_display()
                )
            elif excel_field.title == WORK_HOURS_FIELD_TITLE:
                work_hours = voucher.employment_work_hours
                assert (output_column.value is None and work_hours is None) or Decimal(
                    output_column.value
                ) == work_hours
            elif excel_field.title == SALARY_PAID_FIELD_TITLE:
                salary_paid = voucher.employment_salary_paid
                assert (output_column.value is None and salary_paid is None) or Decimal(
                    output_column.value
                ) == salary_paid
            elif excel_field.title == SUM_FIELD_TITLE:
                assert output_column.value == voucher.value_in_euros
            elif excel_field.model_fields == ["attachments"]:
                expected_attachment_uri = get_attachment_uri(
                    voucher, excel_field, voucher.attachments, response.wsgi_request
                )
                assert output_column.value == expected_attachment_uri
            elif (
                excel_field.title
                in (
                    INVOICER_EMAIL_FIELD_TITLE,
                    INVOICER_NAME_FIELD_TITLE,
                    INVOICER_PHONE_NUMBER_FIELD_TITLE,
                )
                and not voucher.application.is_separate_invoicer
            ):
                assert output_column.value == "", excel_field.title
            elif excel_field.title == VOUCHER_NUMBER_FIELD_TITLE:
                assert output_column.value == voucher.summer_voucher_serial_number
            elif excel_field.model_fields == []:
                assert output_column.value == excel_field.value
            else:
                values_tuple = tuple(
                    handle_special_cases(
                        getattr_nested(voucher, attr_str.split("__")),
                        attr_str,
                        voucher,
                        excel_field,
                        response.wsgi_request,
                    )
                    for attr_str in excel_field.model_fields
                )
                assert output_column.value == excel_field.value % values_tuple, (
                    excel_field.title
                )


@pytest.mark.django_db
@override_settings(NEXT_PUBLIC_MOCK_FLAG=False)
@pytest.mark.parametrize(
    "employer_summer_voucher_creation_date,sum_field_value",
    [
        (date(2021, 6, 1), 325),
        (date(2022, 6, 1), 325),
        (date(2023, 6, 1), 325),
        (date(2024, 1, 1), 325),
        (date(2024, 2, 13), 325),
        (date(2024, 2, 14), 350),
        (date(2024, 12, 31), 350),
    ],
)
def test_excel_view_download_sum_field_value(  # noqa: C901
    staff_client,
    employer_summer_voucher_creation_date: date,
    sum_field_value: int,
):
    with freeze_time(employer_summer_voucher_creation_date):
        EmployerSummerVoucherFactory(
            application=EmployerApplicationFactory(
                status=EmployerApplicationStatus.SUBMITTED
            )
        )

    # Use earlier date because fetching file fails if frozen date is in the future
    with freeze_time(date(employer_summer_voucher_creation_date.year, 1, 1)):
        response = staff_client.get(
            employer_excel_export_url("annual", ExcelColumns.TALPA.value)
        )

    workbook = openpyxl.load_workbook(filename=BytesIO(response.getvalue()))
    rows_generator = workbook.active.rows
    header_row = next(rows_generator)
    data_row = next(rows_generator)

    titles = [column.value for column in header_row]
    sum_field_index = titles.index(SUM_FIELD_TITLE)
    assert data_row[sum_field_index].value == sum_field_value


@pytest.mark.django_db
@override_settings(
    EXCEL_DOWNLOAD_BATCH_SIZE=3,
    NEXT_PUBLIC_MOCK_FLAG=False,
)
def test_youth_excel_download_content(staff_client):  # noqa: C901
    def youth_application_sorting_key(app: YouthApplication):
        # Sorting key should be the same as what is used to order by queryset results
        return app.created_at, app.pk

    InactiveYouthApplicationFactory()  # This should not be exported
    apps: List[YouthApplication] = (
        ActiveYouthApplicationFactory.create_batch(size=12)
        + [
            YouthApplicationFactory(status=status)
            for status in YouthApplicationStatus.active_values()
        ]
        + [
            ActiveVtjTestCaseYouthApplicationFactory(last_name=vtj_test_case)
            for vtj_test_case in VtjTestCase.values
        ]
    )
    # Update created_at specially as its initial save uses auto_now_add
    for app in apps:
        app.created_at -= timedelta(seconds=random.randint(0, 60 * 60 * 24 * 10))
        app.save(update_fields=["created_at"])
        app.refresh_from_db()
    assert min(app.created_at for app in apps) != max(app.created_at for app in apps)
    apps = sorted(apps, key=youth_application_sorting_key)

    with translation.override("en"):  # Should still use Finnish translations
        response = staff_client.get(youth_excel_download_url())

    assert isinstance(response, StreamingHttpResponse)

    workbook = openpyxl.load_workbook(filename=BytesIO(response.getvalue()))
    active_worksheet = workbook.active

    rows_generator = active_worksheet.rows
    header_row = next(rows_generator)
    data_rows = [next(rows_generator) for _ in apps]
    with pytest.raises(StopIteration):
        next(rows_generator)

    output_column_names = [column.value for column in header_row]
    with translation.override("fi"):  # Make extra sure to test for Finnish column names
        expected_output_columns_names = (
            YouthApplicationExcelExportViewSet.output_column_names()
        )
    assert output_column_names == expected_output_columns_names

    source_fields = YouthApplicationExcelExportViewSet.source_fields()
    for data_row, app in zip(data_rows, apps):
        for output_cell, source_field in zip(data_row, source_fields):
            output_value = output_cell.value
            if source_field == "application_date":
                assert date.fromisoformat(output_value) == localdate(app.created_at)
            elif source_field == "id":
                assert output_value == str(app.id)
            elif source_field.endswith("application_year"):
                assert int(output_value) == localdate(app.created_at).year
            elif source_field == "summer_voucher_serial_number":
                if not output_value:
                    assert not app.has_youth_summer_voucher
                else:
                    assert (
                        output_value
                        == app.youth_summer_voucher.user_showable_serial_number
                    )
            elif source_field == "birth_year":
                assert int(output_value) == app.birthdate.year
            elif source_field == "birthdate":
                assert date.fromisoformat(output_value) == app.birthdate
            elif source_field == "is_unlisted_school":
                assert output_value in ["Kyllä", "Ei"]
                assert (output_value == "Kyllä") == app.is_unlisted_school
            elif source_field == "additional_info_user_reasons":
                if output_value == "":
                    assert app.additional_info_user_reasons == []
                else:
                    assert output_value.split(", ") == sorted(
                        app.additional_info_user_reasons
                    )
            elif source_field == "confirmation_date":
                assert date.fromisoformat(output_value) == localdate(
                    app.receipt_confirmed_at
                )
            elif source_field.endswith("additional_info_providing_date"):
                if output_value == "":
                    assert app.additional_info_provided_at is None
                else:
                    assert date.fromisoformat(output_value) == localdate(
                        app.additional_info_provided_at
                    )
            elif source_field == "handling_date":
                if output_value == "":
                    assert app.handled_at is None
                else:
                    assert date.fromisoformat(output_value) == localdate(app.handled_at)
            elif source_field == "target_group":
                from applications.exporters.excel_exporter import (
                    resolve_target_group_and_status,
                )

                expected_val, _ = resolve_target_group_and_status(app)
                assert output_value == expected_val
            elif source_field == "target_group_calculation_status":
                from applications.exporters.excel_exporter import (
                    resolve_target_group_and_status,
                )

                _, expected_status = resolve_target_group_and_status(app)
                assert output_value == expected_status
            else:
                assert output_value == getattr(app, source_field), source_field


@pytest.mark.django_db
@override_settings(NEXT_PUBLIC_MOCK_FLAG=False)
@pytest.mark.parametrize(
    "download_url",
    [
        employer_excel_export_url(download, columns)
        for columns in ExcelColumns.values
        for download in ["unhandled", "annual"]
    ]
    + [youth_excel_download_url()],
)
def test_excel_view_download_with_unauthenticated_user(
    user_client, download_url, accepted_youth_application
):
    create_test_employer_summer_vouchers(year=2021)

    with freeze_time(datetime(2021, 12, 31)):
        response = user_client.get(download_url)

    assert response.status_code == 302
    assert response.url == handler_403_url()


def test_unique_fields_besides_padding_fields():
    field_titles = get_field_titles(FIELDS)
    field_titles_without_padding_fields = [
        field_title for field_title in field_titles if field_title != ""
    ]
    assert len(set(field_titles_without_padding_fields)) == len(
        field_titles_without_padding_fields
    ), "Duplicates in FIELDS besides empty strings"


@pytest.mark.parametrize(
    "field_titles,source",
    [
        (get_field_titles(FIELDS), "FIELDS"),
        (get_field_titles(get_reporting_columns()), "get_reporting_columns"),
        (get_field_titles(get_talpa_columns()), "get_talpa_columns"),
    ],
)
def test_order_field(field_titles, source):
    assert ORDER_FIELD_TITLE in field_titles
    assert field_titles.index(ORDER_FIELD_TITLE) == 0


@pytest.mark.parametrize(
    "field_titles,source",
    [
        (get_field_titles(FIELDS), "FIELDS"),
        (get_field_titles(get_reporting_columns()), "get_reporting_columns"),
        (get_field_titles(get_talpa_columns()), "get_talpa_columns"),
    ],
)
def test_received_data_field(field_titles, source):
    assert RECEIVED_DATE_FIELD_TITLE in field_titles
    assert field_titles.index(RECEIVED_DATE_FIELD_TITLE) == 1


def test_reporting_fields():
    field_titles = get_field_titles(FIELDS)
    reporting_field_titles = get_field_titles(get_reporting_columns())
    assert set(reporting_field_titles) == (
        set(field_titles) - set(REMOVABLE_REPORTING_FIELD_TITLES)
    )
    assert get_reporting_columns() == get_exportable_fields(
        ExcelColumns.REPORTING.value
    )


@override_settings(EXCLUDE_2026_EXCEL_FIELDS=False)
def test_talpa_fields():
    field_titles = get_field_titles(FIELDS)
    talpa_columns = get_talpa_columns()
    talpa_field_titles = get_field_titles(talpa_columns)
    assert set(talpa_field_titles) == (
        set(field_titles) - set(REMOVABLE_TALPA_FIELD_TITLES)
    )
    # The 6 repositioned fields must appear at the end
    end_titles = get_field_titles(talpa_columns[-len(TALPA_END_FIELD_TITLES) :])
    assert end_titles == [str(t) for t in TALPA_END_FIELD_TITLES]
    assert get_talpa_columns() == get_exportable_fields(ExcelColumns.TALPA.value)


@pytest.mark.django_db
@override_settings(EXCLUDE_2026_EXCEL_FIELDS=True)
def test_talpa_fields_with_2026_fields_disabled():
    talpa_columns = get_talpa_columns()
    talpa_field_titles = get_field_titles(talpa_columns)

    for title in TALPA_END_FIELD_TITLES:
        assert str(title) not in talpa_field_titles


def test_removable_reporting_field_titles():
    check_removable_field_titles(REMOVABLE_REPORTING_FIELD_TITLES)


def test_removable_talpa_field_titles():
    check_removable_field_titles(REMOVABLE_TALPA_FIELD_TITLES)


@pytest.mark.django_db
@override_settings(NEXT_PUBLIC_MOCK_FLAG=False)
@pytest.mark.parametrize(
    ("export_url", "expected_error_code"),
    [
        (
            employer_excel_export_url("annual", "invalid-columns"),
            EmployerExcelExportErrorCode.INVALID_COLUMNS,
        ),
        (
            employer_excel_export_url("not-a-real-kind", ExcelColumns.REPORTING.value),
            EmployerExcelExportErrorCode.INVALID_EXPORT_KIND,
        ),
    ],
)
def test_employer_excel_export_rejects_invalid_parameters(
    staff_client, export_url, expected_error_code
):
    """Invalid export paths redirect to the landing page with a stable error code."""
    response = staff_client.get(export_url)

    assert_landing_page_shows_export_error(staff_client, response, expected_error_code)


def test_excel_download_error_messages_are_unique():
    """Each export error code must map to a distinct landing-page message."""
    messages = [
        get_excel_download_error_message(code) for code in EmployerExcelExportErrorCode
    ]
    assert len(messages) == len(set(messages))


@pytest.mark.django_db
@override_settings(NEXT_PUBLIC_MOCK_FLAG=False)
def test_excel_download_landing_page_ignores_unknown_error_code(staff_client):
    """Unknown error query values must not show a landing-page alert."""
    response = staff_client.get(f"{excel_download_url()}?error=not-a-code")

    assert response.status_code == status.HTTP_200_OK
    content = response.content.decode()
    for code in EmployerExcelExportErrorCode:
        assert get_excel_download_error_message(code) not in content


@pytest.mark.django_db
@override_settings(NEXT_PUBLIC_MOCK_FLAG=False)
def test_youth_excel_download_with_missing_birthdate(staff_client):
    """
    Test that exporting youth applications' Excel uses fallback values for
    birth_year and birthdate when birthdate is unavailable (This case should
    not be creatable using end-user UIs, but was spotted e.g. in non-local
    dev-environment and that crashed the youth application Excel export).
    """
    ActiveYouthApplicationFactory(social_security_number="", non_vtj_birthdate=None)

    response = staff_client.get(youth_excel_download_url())

    assert isinstance(response, StreamingHttpResponse)
    assert response.status_code == 200

    workbook = openpyxl.load_workbook(filename=BytesIO(response.getvalue()))
    rows_generator = workbook.active.rows
    _header_row = next(rows_generator)
    data_row = next(rows_generator)  # exactly one application was created
    with pytest.raises(StopIteration):
        next(rows_generator)

    source_fields = YouthApplicationExcelExportViewSet.source_fields()
    row_by_field = dict(zip(source_fields, (cell.value for cell in data_row)))

    # Test that birth_year and birthdate use their respective fallback values:
    assert row_by_field["birth_year"] == -1
    assert row_by_field["birthdate"] == "N/A"


@pytest.mark.django_db
@override_settings(NEXT_PUBLIC_MOCK_FLAG=True)
def test_youth_excel_download_unauthenticated_redirects_with_mock_flag(client):
    """
    Test that when mock flag is on, a GET request on youth-excel-download
    by a plain unauthenticated client redirects to the ADFS login page.
    """
    response = client.get(youth_excel_download_url())
    assert response.status_code == status.HTTP_302_FOUND
    assert response.url.startswith(reverse("django_auth_adfs:login"))


@override_settings(NEXT_PUBLIC_MOCK_FLAG=False)
def test_employer_excel_target_group_calculated_when_missing(staff_client):
    """When youth application has no target_group, Excel should calculate it from age."""
    voucher = EmployerSummerVoucherFactory(
        youth_summer_voucher__youth_application__social_security_number="",
        youth_summer_voucher__youth_application__non_vtj_birthdate=date(
            date.today().year - 16, 7, 1
        ),
    )
    youth_app = voucher.youth_summer_voucher.youth_application
    youth_app.target_group = ""
    youth_app.save(update_fields=["target_group"])

    response = staff_client.get(
        employer_excel_export_url("annual", ExcelColumns.REPORTING.value)
    )
    workbook = openpyxl.load_workbook(filename=BytesIO(response.getvalue()))
    rows = list(workbook.active.rows)
    header = [c.value for c in rows[0]]
    data = [c.value for c in rows[1]]

    with translation.override("fi"):
        special_case_title = str(SPECIAL_CASE_FIELD_TITLE)
    idx = header.index(special_case_title)
    # Should have calculated a target group from the youth's age
    assert data[idx] != "" and data[idx] is not None


@pytest.mark.django_db
@override_settings(NEXT_PUBLIC_MOCK_FLAG=False)
def test_employer_excel_target_group_unresolved_with_underage_or_overage(staff_client):
    """When youth has underage_or_overage and no target_group, Excel should leave it empty."""
    voucher = EmployerSummerVoucherFactory(
        youth_summer_voucher__youth_application__social_security_number="",
        youth_summer_voucher__youth_application__non_vtj_birthdate=date(
            date.today().year - 16, 7, 1
        ),
    )
    youth_app = voucher.youth_summer_voucher.youth_application
    youth_app.target_group = ""
    youth_app.additional_info_user_reasons = [
        AdditionalInfoUserReason.UNDERAGE_OR_OVERAGE.value
    ]
    youth_app.save(update_fields=["target_group", "additional_info_user_reasons"])

    response = staff_client.get(
        employer_excel_export_url("annual", ExcelColumns.REPORTING.value)
    )
    workbook = openpyxl.load_workbook(filename=BytesIO(response.getvalue()))
    rows = list(workbook.active.rows)
    header = [c.value for c in rows[0]]
    data = [c.value for c in rows[1]]

    with translation.override("fi"):
        special_case_title = str(SPECIAL_CASE_FIELD_TITLE)
    idx = header.index(special_case_title)
    assert data[idx] in ("", None)


@pytest.mark.django_db
@override_settings(NEXT_PUBLIC_MOCK_FLAG=False)
@pytest.mark.parametrize(
    "target_group,additional_reasons,expected_status",
    [
        ("primary_target_group", [], "annettu"),
        ("", [], "laskettu"),
        ("", [AdditionalInfoUserReason.UNDERAGE_OR_OVERAGE.value], "selvittämätön"),
    ],
)
def test_employer_excel_calculation_status_values(
    staff_client, target_group, additional_reasons, expected_status
):
    """Calculation status field shows correct value based on target group state."""
    voucher = EmployerSummerVoucherFactory(
        youth_summer_voucher__youth_application__social_security_number="",
        youth_summer_voucher__youth_application__non_vtj_birthdate=date(
            date.today().year - 16, 7, 1
        ),
    )
    youth_app = voucher.youth_summer_voucher.youth_application
    youth_app.target_group = target_group
    youth_app.additional_info_user_reasons = additional_reasons
    youth_app.save(update_fields=["target_group", "additional_info_user_reasons"])

    response = staff_client.get(
        employer_excel_export_url("annual", ExcelColumns.REPORTING.value)
    )
    workbook = openpyxl.load_workbook(filename=BytesIO(response.getvalue()))
    rows = list(workbook.active.rows)
    header = [c.value for c in rows[0]]
    data = [c.value for c in rows[1]]

    with translation.override("fi"):
        status_title = str(_("Erikoistapauksen laskentatila"))
    idx = header.index(status_title)
    assert data[idx] == expected_status


@pytest.mark.django_db
def test_resolve_target_group_and_status_always_finnish():
    from applications.exporters.excel_exporter import resolve_target_group_and_status

    # 1. No youth application
    with translation.override("en"):
        _, status1 = resolve_target_group_and_status(None)
        assert status1 == "selvittämätön"

    with translation.override("sv"):
        _, status2 = resolve_target_group_and_status(None)
        assert status2 == "selvittämätön"

    # 2. Youth application with target group
    app = YouthApplicationFactory(target_group="primary_target_group")
    with translation.override("en"):
        display, status = resolve_target_group_and_status(app)
        assert display == "9. luokkalainen"
        assert status == "annettu"

    with translation.override("sv"):
        display, status = resolve_target_group_and_status(app)
        assert display == "9. luokkalainen"
        assert status == "annettu"

    # 3. Youth application with underage/overage reason
    app_err = YouthApplicationFactory(
        target_group="",
        additional_info_user_reasons=[
            AdditionalInfoUserReason.UNDERAGE_OR_OVERAGE.value
        ],
    )
    with translation.override("en"):
        display, status = resolve_target_group_and_status(app_err)
        assert display == ""
        assert status == "selvittämätön"

    # 4. Youth application where target group class is resolved by age (e.g. age 15)
    app_calculated = YouthApplicationFactory(
        target_group="",
        social_security_number="",
        non_vtj_birthdate=date(date.today().year - 15, 7, 1),
    )
    with translation.override("en"):
        display, status = resolve_target_group_and_status(app_calculated)
        assert display == "8. luokkalainen"
        assert status == "laskettu"


@pytest.mark.django_db
@override_settings(NEXT_PUBLIC_MOCK_FLAG=False)
@pytest.mark.parametrize("exclude_flag", [False, True])
def test_talpa_excel_gated_2026_and_status_fields(staff_client, settings, exclude_flag):
    """The 2026 fields and status column should be present or absent based on EXCLUDE_2026_EXCEL_FIELDS."""
    settings.EXCLUDE_2026_EXCEL_FIELDS = exclude_flag

    EmployerSummerVoucherFactory(
        application=EmployerApplicationFactory(
            status=EmployerApplicationStatus.SUBMITTED
        )
    )

    response = staff_client.get(
        employer_excel_export_url("annual", ExcelColumns.TALPA.value)
    )
    workbook = openpyxl.load_workbook(filename=BytesIO(response.getvalue()))
    header = [c.value for c in next(workbook.active.rows)]

    expected_end = [
        "VTJ-tietojen luovutuskielto (ts. turvakielto)",
        "Maksunsaajan nimi",
        "Maksunsaajan osoite",
        "Pankin SWIFT / BIC koodi",
        "Pankin nimi",
        "Pankin käyntiosoite",
        "Erikoistapauksen laskentatila",
    ]

    should_include = not exclude_flag
    if should_include:
        assert header[-len(expected_end) :] == expected_end
    else:
        for title in expected_end:
            assert title not in header
